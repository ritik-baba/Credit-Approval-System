import openpyxl
from credit_api.models import Customer, Loan
from background_task import background
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import Customer, Loan
from .serializers import CustomerSerializer, LoanSerializer
from django.db.models import Sum
from django.db.models import Max
from .models import Customer
from .serializers import CustomerSerializer
import openpyxl
from credit_api.models import Customer, Loan
from background_task import background
from django.utils.dateparse import parse_date
from datetime import datetime


@background(schedule=1)
def ingest_customer_data():
    workbook = openpyxl.load_workbook('customer_data.xlsx')
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not all([row[0], row[1], row[2], row[3], row[4], row[5], row[6]]):
            print("Skipping row due to missing data:", row)
            continue

        # Check if the customer already exists
        if Customer.objects.filter(customer_id=row[0]).exists():
            print(f"Customer with ID {row[0]} already exists. Skipping.")
            continue

        # Create the customer if not already present
        Customer.objects.create(
            customer_id=row[0],
            first_name=row[1],
            last_name=row[2],
            phone_number=row[4],
            monthly_salary=row[5],
            approved_limit=row[6],
            current_debt=0  # Default to 0 or adjust based on your logic
        )

@background(schedule=1)
def ingest_loan_data():
    workbook = openpyxl.load_workbook('loan_data.xlsx')
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(f"Processing row: {row}")  # Debug: Print the entire row
        try:
            # Check if the customer exists
            customer = Customer.objects.get(customer_id=row[0])
            print(f"Found customer: {customer}")  # Debug: Print customer

            # Convert datetime to string if necessary
            start_date = row[7]
            end_date = row[8]
            if isinstance(start_date, datetime):
                start_date = start_date.strftime("%Y-%m-%d")
            if isinstance(end_date, datetime):
                end_date = end_date.strftime("%Y-%m-%d")

            # Parse dates
            start_date = parse_date(start_date) if start_date else None
            end_date = parse_date(end_date) if end_date else None

            # Check for None values in required fields
            if not all([row[2], row[3], row[4], row[5], row[6]]):
                print(f"Skipping row due to missing values: {row}")
                continue

            # Create Loan object
            Loan.objects.create(
                customer=customer,
                loan_amount=row[2],
                tenure=row[3],
                interest_rate=row[4],
                monthly_repayment=row[5],
                emis_paid_on_time=row[6],
                start_date=start_date,
                end_date=end_date
            )
            print(f"Loan created for customer ID {row[0]}")  # Debug: Confirm loan creation
        except Customer.DoesNotExist:
            print(f"Customer with ID {row[0]} does not exist. Skipping this loan.")
        except Exception as e:
            print(f"Error processing loan for Customer ID {row[0]}: {e}")


@api_view(['POST'])
def register_customer(request):
    data = request.data
    approved_limit = round(36 * data['monthly_income'] / 100000) * 100000

    # Get the last customer_id and increment it
    last_customer = Customer.objects.aggregate(Max('customer_id'))
    last_customer_id = last_customer['customer_id__max'] or 0
    new_customer_id = last_customer_id + 1

    # Create the new customer with the incremented customer_id
    customer = Customer.objects.create(
        customer_id=new_customer_id,
        first_name=data['first_name'],
        last_name=data['last_name'],
        age=data['age'],
        monthly_salary=data['monthly_income'],
        phone_number=data['phone_number'],
        approved_limit=approved_limit
    )

    serializer = CustomerSerializer(customer)
    return Response(serializer.data)

@api_view(['POST'])
def check_eligibility(request):
    data = request.data
    customer_id = data['customer_id']
    loan_amount = data['loan_amount']
    interest_rate = data['interest_rate']
    tenure = data['tenure']

    try:
        customer = Customer.objects.get(customer_id=customer_id)
        total_debt = Loan.objects.filter(customer=customer).aggregate(Sum('loan_amount'))['loan_amount__sum'] or 0
        if total_debt > customer.approved_limit:
            return Response({
                "customer_id": customer_id,
                "approval": False,
                "interest_rate": interest_rate,
                "corrected_interest_rate": None,
                "tenure": tenure,
                "monthly_installment": None
            })

        # Calculate credit score
        loans = Loan.objects.filter(customer=customer)
        loans_paid_on_time = sum(loan.emis_paid_on_time for loan in loans)
        number_of_loans = loans.count()
        loan_activity_this_year = loans.filter(start_date__year=datetime.now().year).count()

        credit_score = min(100, loans_paid_on_time - number_of_loans + loan_activity_this_year)
        if total_debt > 0.5 * customer.monthly_salary:
            credit_score = 0

        # Determine interest rate slab
        corrected_interest_rate = interest_rate
        if credit_score > 50:
            approved = True
        elif 30 < credit_score <= 50:
            approved = True
            corrected_interest_rate = max(interest_rate, 12)
        elif 10 < credit_score <= 30:
            approved = True
            corrected_interest_rate = max(interest_rate, 16)
        else:
            approved = False

        if approved and (loan_amount + total_debt) > customer.approved_limit:
            approved = False

        # Calculate monthly installment using compound interest formula
        if approved:
            r = corrected_interest_rate / (12 * 100)
            monthly_installment = loan_amount * r * (1 + r) ** tenure / ((1 + r) ** tenure - 1)
            monthly_installment = round(monthly_installment, 2)
        else:
            monthly_installment = None

        return Response({
            "customer_id": customer_id,
            "approval": approved,
            "interest_rate": interest_rate,
            "corrected_interest_rate": corrected_interest_rate,
            "tenure": tenure,
            "monthly_installment": monthly_installment
        })
    except Customer.DoesNotExist:
        return Response({"error": "Customer not found"}, status=404)

@api_view(['POST'])
def create_loan(request):
    data = request.data
    customer_id = data['customer_id']
    loan_amount = data['loan_amount']
    interest_rate = data['interest_rate']
    tenure = data['tenure']

    try:
        customer = Customer.objects.get(customer_id=customer_id)
        total_debt = Loan.objects.filter(customer=customer).aggregate(Sum('loan_amount'))['loan_amount__sum'] or 0

        if total_debt + loan_amount > customer.approved_limit:
            return Response({
                "loan_id": None,
                "customer_id": customer_id,
                "loan_approved": False,
                "message": "Loan amount exceeds approved credit limit",
                "monthly_installment": None
            })

        # Calculate monthly installment
        r = interest_rate / (12 * 100)
        monthly_installment = loan_amount * r * (1 + r) ** tenure / ((1 + r) ** tenure - 1)
        monthly_installment = round(monthly_installment, 2)

        loan = Loan.objects.create(
            customer=customer,
            loan_amount=loan_amount,
            tenure=tenure,
            interest_rate=interest_rate,
            monthly_repayment=monthly_installment,
            emis_paid_on_time=0,
            start_date=datetime.now().date(),
            end_date=datetime.now().date().replace(year=datetime.now().year + tenure // 12)
        )

        return Response({
            "loan_id": loan.loan_id,
            "customer_id": customer_id,
            "loan_approved": True,
            "message": "Loan approved successfully",
            "monthly_installment": monthly_installment
        })
    except Customer.DoesNotExist:
        return Response({"error": "Customer not found"}, status=404)

@api_view(['GET'])
def view_loan(request, loan_id):
    try:
        loan = Loan.objects.get(loan_id=loan_id)
        serializer = LoanSerializer(loan)
        return Response(serializer.data)
    except Loan.DoesNotExist:
        return Response({"error": "Loan not found"}, status=404)

@api_view(['GET'])
def view_loans_by_customer(request, customer_id):
    loans = Loan.objects.filter(customer__customer_id=customer_id)
    serializer = LoanSerializer(loans, many=True)
    return Response(serializer.data)

from django.http import HttpResponse

def home(request):
    return HttpResponse("Welcome to the Credit Approval System API!")
